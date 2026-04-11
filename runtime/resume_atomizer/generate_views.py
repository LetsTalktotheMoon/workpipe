#!/usr/bin/env python3
"""
generate_views.py — Phase 5: 从 atom_store.json + skill_tree.json 生成所有视图文件。

输出:
- master_resume.md   (视角1: 正向，像简历一样阅读)
- skill_index.md     (视角2: 反向，按技能树查找要点)
- coverage_matrix.tsv (技能×简历矩阵)
- resume_analysis.xlsx (6 Sheet Excel 工作簿)

Usage:
    python3 generate_views.py
"""

import json
import re
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).parent
ATOM_STORE = ROOT / "atom_store.json"
SKILL_TREE = ROOT / "skill_tree.json"


def load_data():
    store = json.loads(ATOM_STORE.read_text(encoding='utf-8'))
    tree = json.loads(SKILL_TREE.read_text(encoding='utf-8'))
    # 建 atom 查找表
    atom_map = {a['atom_id']: a for a in store['atoms']}
    return store, tree, atom_map


# ========================
# 视角1: master_resume.md
# ========================
def generate_master_resume(store, tree, atom_map):
    lines = []
    lines.append("# Master Resume\n")
    lines.append(f"*Generated from {store['_meta']['source_resume_count']} tailored resumes, "
                 f"{store['_meta']['atom_count']} atoms*\n")
    lines.append("---\n")

    # === 万能要点速查索引 ===
    universal_atoms = [a for a in store['atoms'] if a.get('universal')]
    if universal_atoms:
        lines.append("## 📌 万能要点索引 (Universal Atoms)\n")
        lines.append(f"*以下 {len(universal_atoms)} 个要点出现在 10+ 份简历中，适用于多种岗位方向*\n")
        for a in universal_atoms:
            exp_name = _get_exp_name(store, a['parent_exp_id'])
            skills_str = ', '.join(a['skill_ids'][:5])
            lines.append(f"- **{a['atom_id']}** ({exp_name}): {a['text'][:80]}... "
                         f"`[{skills_str}]`\n")
        lines.append("\n---\n")

    # === Summary ===
    lines.append("## Professional Summary\n")
    lines.append("*每个 slot 有 22 个版本，按目标岗位选取*\n")
    for sa in store['summary_atoms']:
        lines.append(f"\n### Slot {sa['position']}: {sa['role']}\n")
        for v in sa['variants']:
            skills_str = ', '.join(v.get('skill_ids', [])[:5])
            rid = v['rid']
            lines.append(f"- **[R{rid:02d}]** {v['text'][:200]}{'...' if len(v['text']) > 200 else ''}\n")
            if skills_str:
                lines.append(f"  `skills: [{skills_str}]`\n")

    # === Work Experience ===
    for exp in store['experiences']:
        lines.append(f"\n---\n")
        lines.append(f"## {exp['title']} | {exp['company']} · {exp['division']} | {exp['location']}\n")
        lines.append(f"**{exp['period']}**\n")

        for g in exp['contribution_groups']:
            gtype = g['type']
            pick = g.get('pick_mode', 'pick_all')
            n_atoms = len(g['atom_ids'])

            if gtype == 'achievement_cluster':
                lines.append(f"\n### 🔄 {g['canonical_title']} *(pick_one, {n_atoms} variants from R{g['source_resumes']})*\n")
            elif gtype == 'named_project':
                lines.append(f"\n### 📋 {g['canonical_title']}\n")
            elif gtype == 'core_contributions':
                lines.append(f"\n### ⚙️ {g['canonical_title']}\n")
            elif gtype == 'themed_section':
                lines.append(f"\n### 🏷️ {g['canonical_title']}\n")
            else:
                lines.append(f"\n### {g['canonical_title']}\n")

            for aid in g['atom_ids']:
                atom = atom_map.get(aid)
                if not atom:
                    continue
                u_tag = " **[U]**" if atom.get('universal') else ""
                skills_short = ', '.join(s.split('.')[-1] for s in atom['skill_ids'][:6])
                angles = ', '.join(atom.get('angle_tags', []))
                lines.append(f"- {u_tag} *[{aid}, R{atom['source_resumes'][0]:02d}]* {atom['text']}\n")
                lines.append(f"  `skills: [{skills_short}]` `angles: [{angles}]`\n")

    # === Academic Projects ===
    if store['academic_projects']:
        lines.append("\n---\n")
        lines.append("## Academic Projects\n")
        for proj in store['academic_projects']:
            lines.append(f"\n### {proj['canonical_title']} — {proj['course']}\n")
            for aid in proj['atom_ids']:
                atom = atom_map.get(aid)
                if atom:
                    skills_short = ', '.join(s.split('.')[-1] for s in atom['skill_ids'][:6])
                    lines.append(f"- *[{aid}]* {atom['text']}\n")
                    lines.append(f"  `skills: [{skills_short}]`\n")

    # === Education ===
    lines.append("\n---\n")
    lines.append("## Education\n")
    for edu in store['identity']['education']:
        lines.append(f"- **{edu['degree']}** | {edu['institution']} | {edu['period']}\n")

    # === Additional ===
    lines.append("\n---\n")
    lines.append("## Additional Information\n")
    for add in store['identity']['additional']:
        lines.append(f"- **{add['category'].title()}:** {add['text']}\n")

    output = '\n'.join(lines)
    (ROOT / "master_resume.md").write_text(output, encoding='utf-8')
    print(f"✅ master_resume.md: {len(lines)} lines")


# ========================
# 视角2: skill_index.md
# ========================
def generate_skill_index(store, tree, atom_map):
    lines = []
    lines.append("# Skill Index — 按技能查找要点\n")
    lines.append(f"*{store['_meta']['atom_count']} atoms mapped to skill tree*\n")
    lines.append("---\n")

    # 建反向索引: skill_id → [atom_ids]
    skill_to_atoms = defaultdict(list)
    for a in store['atoms']:
        for sid in a['skill_ids']:
            skill_to_atoms[sid].append(a['atom_id'])

    # Hard skills
    lines.append("## Hard Skills\n")
    for domain in tree['hard_skill_tree']:
        lines.append(f"\n### {domain['domain']}\n")
        for sd in domain['sub_domains']:
            lines.append(f"\n#### {sd['name']}\n")
            for skill in sd['skills']:
                skill_id = f"hard.{_slug(domain['domain'])}.{_slug(sd['name'])}.{_slug(skill)}"
                atoms = skill_to_atoms.get(skill_id, [])
                count = len(atoms)
                if count == 0:
                    lines.append(f"- **{skill}** — *(无对应要点)*\n")
                else:
                    lines.append(f"- **{skill}** ({count} atoms)\n")
                    for aid in atoms[:10]:  # 最多显示10个
                        atom = atom_map.get(aid)
                        if atom:
                            exp_name = _get_exp_short(store, atom['parent_exp_id'])
                            u = "[U] " if atom.get('universal') else ""
                            lines.append(f"  - {u}`{aid}` ({exp_name}, R{atom['source_resumes'][0]:02d}): "
                                         f"{atom['text'][:100]}...\n")
                    if count > 10:
                        lines.append(f"  - *... +{count - 10} more*\n")

    # Soft skills
    lines.append("\n---\n")
    lines.append("## Soft Skills\n")
    for cluster in tree['soft_skill_clusters']:
        lines.append(f"\n### {cluster['cluster_name']}\n")
        for skill in cluster['skills']:
            skill_id = f"soft.{_slug(cluster['cluster_name'])}.{_slug(skill)}"
            atoms = skill_to_atoms.get(skill_id, [])
            count = len(atoms)
            if count == 0:
                lines.append(f"- **{skill}** — *(无对应要点)*\n")
            else:
                lines.append(f"- **{skill}** ({count} atoms)\n")
                for aid in atoms[:5]:
                    atom = atom_map.get(aid)
                    if atom:
                        exp_name = _get_exp_short(store, atom['parent_exp_id'])
                        lines.append(f"  - `{aid}` ({exp_name}): {atom['text'][:80]}...\n")
                if count > 5:
                    lines.append(f"  - *... +{count - 5} more*\n")

    output = '\n'.join(lines)
    (ROOT / "skill_index.md").write_text(output, encoding='utf-8')
    print(f"✅ skill_index.md: {len(lines)} lines")


# ========================
# coverage_matrix.tsv
# ========================
def generate_coverage_matrix(store, tree, atom_map):
    # 建 skill → resume → count 矩阵
    matrix = defaultdict(lambda: defaultdict(int))
    all_rids = sorted(set(
        rid for a in store['atoms'] for rid in a['source_resumes']
    ))

    for a in store['atoms']:
        for sid in a['skill_ids']:
            for rid in a['source_resumes']:
                matrix[sid][rid] += 1

    # 收集所有 skill_ids（按 tree 顺序）
    ordered_skills = []
    for domain in tree['hard_skill_tree']:
        for sd in domain['sub_domains']:
            for skill in sd['skills']:
                sid = f"hard.{_slug(domain['domain'])}.{_slug(sd['name'])}.{_slug(skill)}"
                ordered_skills.append((sid, skill, domain['domain'], sd['name']))
    for cluster in tree['soft_skill_clusters']:
        for skill in cluster['skills']:
            sid = f"soft.{_slug(cluster['cluster_name'])}.{_slug(skill)}"
            ordered_skills.append((sid, skill, cluster['cluster_name'], ''))

    # 写 TSV
    header = ['skill_id', 'skill_name', 'domain', 'subdomain'] + [f'R{r:02d}' for r in all_rids]
    rows = ['\t'.join(header)]
    for sid, name, domain, subdomain in ordered_skills:
        row = [sid, name, domain, subdomain]
        for rid in all_rids:
            row.append(str(matrix[sid].get(rid, 0)))
        rows.append('\t'.join(row))

    (ROOT / "coverage_matrix.tsv").write_text('\n'.join(rows), encoding='utf-8')
    print(f"✅ coverage_matrix.tsv: {len(ordered_skills)} skills × {len(all_rids)} resumes")


# ========================
# resume_analysis.xlsx
# ========================
def generate_excel(store, tree, atom_map):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
    except ImportError:
        print("⚠️ openpyxl 未安装，跳过 Excel 生成")
        return

    wb = Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    def write_header(ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
        ws.auto_filter.ref = ws.dimensions

    # === Sheet 1: Atoms_Master ===
    ws1 = wb.active
    ws1.title = "Atoms_Master"
    headers = ["atom_id", "parent_exp", "parent_group", "group_type", "pick_mode",
               "canonical_text", "skill_tags", "source_resumes", "universal", "angle_tags"]
    write_header(ws1, headers)
    for i, a in enumerate(store['atoms'], 2):
        # 找 group info
        group_title = ""
        group_type = ""
        pick_mode = ""
        for exp in store['experiences']:
            for g in exp['contribution_groups']:
                if a['atom_id'] in g['atom_ids']:
                    group_title = g['canonical_title']
                    group_type = g['type']
                    pick_mode = g.get('pick_mode', '')
                    break
        for proj in store.get('academic_projects', []):
            if a['atom_id'] in proj.get('atom_ids', []):
                group_title = proj['canonical_title']
                group_type = 'academic_project'
                break

        ws1.cell(row=i, column=1, value=a['atom_id'])
        ws1.cell(row=i, column=2, value=a['parent_exp_id'])
        ws1.cell(row=i, column=3, value=group_title)
        ws1.cell(row=i, column=4, value=group_type)
        ws1.cell(row=i, column=5, value=pick_mode)
        ws1.cell(row=i, column=6, value=a['text'][:500])
        ws1.cell(row=i, column=7, value=', '.join(a['skill_ids']))
        ws1.cell(row=i, column=8, value=', '.join(str(r) for r in a['source_resumes']))
        ws1.cell(row=i, column=9, value='Y' if a.get('universal') else '')
        ws1.cell(row=i, column=10, value=', '.join(a.get('angle_tags', [])))

    # 设置列宽
    for col, width in [(1, 10), (2, 15), (3, 35), (4, 18), (5, 10),
                        (6, 80), (7, 50), (8, 15), (9, 8), (10, 20)]:
        ws1.column_dimensions[chr(64 + col)].width = width

    # === Sheet 2: Atom_Skill_Long ===
    ws2 = wb.create_sheet("Atom_Skill_Long")
    headers = ["atom_id", "skill_id", "skill_name", "skill_type", "domain", "subdomain",
               "parent_exp", "source_resume", "universal"]
    write_header(ws2, headers)
    row = 2
    skill_info = _build_skill_info(tree)
    for a in store['atoms']:
        for sid in a['skill_ids']:
            info = skill_info.get(sid, {})
            ws2.cell(row=row, column=1, value=a['atom_id'])
            ws2.cell(row=row, column=2, value=sid)
            ws2.cell(row=row, column=3, value=info.get('name', sid))
            ws2.cell(row=row, column=4, value=info.get('type', ''))
            ws2.cell(row=row, column=5, value=info.get('domain', ''))
            ws2.cell(row=row, column=6, value=info.get('subdomain', ''))
            ws2.cell(row=row, column=7, value=a['parent_exp_id'])
            ws2.cell(row=row, column=8, value=a['source_resumes'][0] if a['source_resumes'] else '')
            ws2.cell(row=row, column=9, value='Y' if a.get('universal') else '')
            row += 1
    print(f"  Atom_Skill_Long: {row - 2} rows")

    # === Sheet 3: Dim_Skill_Tree ===
    ws3 = wb.create_sheet("Dim_Skill_Tree")
    headers = ["skill_id", "skill_name", "skill_type", "domain", "subdomain", "atom_count"]
    write_header(ws3, headers)
    row = 2
    # 计算每个 skill 的 atom 数
    skill_atom_count = defaultdict(int)
    for a in store['atoms']:
        for sid in a['skill_ids']:
            skill_atom_count[sid] += 1

    for sid, info in sorted(skill_info.items()):
        ws3.cell(row=row, column=1, value=sid)
        ws3.cell(row=row, column=2, value=info['name'])
        ws3.cell(row=row, column=3, value=info['type'])
        ws3.cell(row=row, column=4, value=info['domain'])
        ws3.cell(row=row, column=5, value=info['subdomain'])
        ws3.cell(row=row, column=6, value=skill_atom_count.get(sid, 0))
        row += 1

    # === Sheet 4: Exp_Groups ===
    ws4 = wb.create_sheet("Exp_Groups")
    headers = ["exp_id", "company", "group_id", "group_type", "pick_mode",
               "canonical_title", "atom_count", "source_resumes"]
    write_header(ws4, headers)
    row = 2
    for exp in store['experiences']:
        for g in exp['contribution_groups']:
            ws4.cell(row=row, column=1, value=exp['exp_id'])
            ws4.cell(row=row, column=2, value=exp['company'])
            ws4.cell(row=row, column=3, value=g['group_id'])
            ws4.cell(row=row, column=4, value=g['type'])
            ws4.cell(row=row, column=5, value=g.get('pick_mode', ''))
            ws4.cell(row=row, column=6, value=g['canonical_title'])
            ws4.cell(row=row, column=7, value=len(g['atom_ids']))
            ws4.cell(row=row, column=8, value=', '.join(str(r) for r in g.get('source_resumes', [])))
            row += 1

    # === Sheet 5: Coverage_Matrix ===
    ws5 = wb.create_sheet("Coverage_Matrix")
    all_rids = sorted(set(rid for a in store['atoms'] for rid in a['source_resumes']))
    matrix = defaultdict(lambda: defaultdict(int))
    for a in store['atoms']:
        for sid in a['skill_ids']:
            for rid in a['source_resumes']:
                matrix[sid][rid] += 1

    headers = ["skill_name", "domain", "subdomain"] + [f'R{r:02d}' for r in all_rids]
    write_header(ws5, headers)
    row = 2
    for sid, info in sorted(skill_info.items()):
        ws5.cell(row=row, column=1, value=info['name'])
        ws5.cell(row=row, column=2, value=info['domain'])
        ws5.cell(row=row, column=3, value=info['subdomain'])
        for ci, rid in enumerate(all_rids, 4):
            val = matrix[sid].get(rid, 0)
            if val > 0:
                ws5.cell(row=row, column=ci, value=val)
        row += 1

    # === Sheet 6: Summary_Stats ===
    ws6 = wb.create_sheet("Summary_Stats")
    headers = ["resume_id", "label", "total_atoms", "unique_skills", "universal_atoms", "angle_coverage"]
    write_header(ws6, headers)
    for ri, rid in enumerate(all_rids, 2):
        atoms_for_resume = [a for a in store['atoms'] if rid in a['source_resumes']]
        skills_set = set(s for a in atoms_for_resume for s in a['skill_ids'])
        universal_count = sum(1 for a in atoms_for_resume if a.get('universal'))
        angles = set(t for a in atoms_for_resume for t in a.get('angle_tags', []))
        ws6.cell(row=ri, column=1, value=rid)
        ws6.cell(row=ri, column=2, value=f"R{rid:02d}")
        ws6.cell(row=ri, column=3, value=len(atoms_for_resume))
        ws6.cell(row=ri, column=4, value=len(skills_set))
        ws6.cell(row=ri, column=5, value=universal_count)
        ws6.cell(row=ri, column=6, value=', '.join(sorted(angles)))

    # 保存
    output_path = ROOT / "resume_analysis.xlsx"
    wb.save(output_path)
    print(f"✅ resume_analysis.xlsx: 6 sheets")


def _get_exp_name(store, exp_id):
    for exp in store['experiences']:
        if exp['exp_id'] == exp_id:
            return exp['company']
    return exp_id


def _get_exp_short(store, exp_id):
    mapping = {"exp-bytedance": "BT", "exp-didi": "DD", "exp-temu": "TM", "acad": "ACAD"}
    return mapping.get(exp_id, exp_id)


def _build_skill_info(tree):
    info = {}
    for domain in tree['hard_skill_tree']:
        for sd in domain['sub_domains']:
            for skill in sd['skills']:
                sid = f"hard.{_slug(domain['domain'])}.{_slug(sd['name'])}.{_slug(skill)}"
                info[sid] = {
                    'name': skill, 'type': 'hard',
                    'domain': domain['domain'], 'subdomain': sd['name']
                }
    for cluster in tree['soft_skill_clusters']:
        for skill in cluster['skills']:
            sid = f"soft.{_slug(cluster['cluster_name'])}.{_slug(skill)}"
            info[sid] = {
                'name': skill, 'type': 'soft',
                'domain': cluster['cluster_name'], 'subdomain': ''
            }
    return info


def _slug(text):
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')[:30]


def main():
    store, tree, atom_map = load_data()
    generate_master_resume(store, tree, atom_map)
    generate_skill_index(store, tree, atom_map)
    generate_coverage_matrix(store, tree, atom_map)
    generate_excel(store, tree, atom_map)
    print(f"\n🎉 全部视图生成完成")


if __name__ == "__main__":
    main()
